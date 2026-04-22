#!/usr/bin/env python3
"""Anonymize personal data in Scoreg-like Excel exports.

What gets anonymized:
- Names (first/middle/last/full-name fields)
- Street names in address fields
- Email addresses
- Phone numbers
- Birth dates and entry dates
- Postal code (PLZ) and country
- IDs such as 1-GGG-X12123, including embedded IDs in text

A mapping JSON is persisted so repeated runs keep replacements stable
across different input files.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime
import json
import random
import re
import string
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Set

from openpyxl import load_workbook

# First names from Star Trek and The Simpsons universes.
FIRST_NAME_POOL: List[str] = [
    "James",
    "JeanLuc",
    "Benjamin",
    "Kathryn",
    "Leonard",
    "Spock",
    "Nyota",
    "Montgomery",
    "Hikaru",
    "Pavel",
    "Deanna",
    "William",
    "Data",
    "Geordi",
    "Beverly",
    "Worf",
    "Julian",
    "Jadzia",
    "Ezri",
    "Kira",
    "Quark",
    "Odo",
    "Rom",
    "Nog",
    "Miles",
    "Keiko",
    "Tom",
    "BElanna",
    "Harry",
    "Tuvok",
    "Neelix",
    "Chakotay",
    "Seven",
    "Phlox",
    "Hoshi",
    "Travis",
    "Malcolm",
    "Michael",
    "Saru",
    "Sylvia",
    "Homer",
    "Marge",
    "Bart",
    "Lisa",
    "Maggie",
    "Ned",
    "Maude",
    "Milhouse",
    "Nelson",
    "Ralph",
    "Martin",
    "Apu",
    "Moe",
    "Lenny",
    "Carl",
    "Barney",
    "Krusty",
    "Sideshow",
    "Montgomery",
    "Waylon",
    "Seymour",
    "Edna",
    "Patty",
    "Selma",
    "Abraham",
    "Mona",
    "Clancy",
    "Luann",
    "Helen",
    "Rainier",
    "Agnes",
    "Sherri",
    "Terri",
    "Snake",
    "Chief",
    "Wiggum",
    "Comic",
    "Manjula",
    "Rod",
    "Todd",
    "Otto",
    "Troy",
    "Kent",
    "Lurleen",
    "Jacqueline",
    "Ling",
    "Akira",
    "Sela",
    "Ro",
    "Keiko",
    "Kes",
    "Icheb",
    "Naomi",
    "Kestra",
    "Boimler",
    "Mariner",
    "Tendi",
    "Rutherford",
    "Freeman",
    "Shaxs",
    "Kayshon",
    "TPol",
]

LAST_NAME_POOL: List[str] = [
    "Kirk",
    "Picard",
    "Sisko",
    "Janeway",
    "McCoy",
    "Spock",
    "Uhura",
    "Scotty",
    "Sulu",
    "Chekov",
    "Troi",
    "Riker",
    "Crusher",
    "LaForge",
    "Worf",
    "Bashir",
    "Dax",
    "Kira",
    "OBrien",
    "Paris",
    "Torres",
    "Kim",
    "Tuvok",
    "Neelix",
    "Archer",
    "Sato",
    "Mayweather",
    "Reed",
    "Burnham",
    "Saru",
    "Georgiou",
    "Pike",
    "Spiner",
    "Q",
    "Borg",
    "Simpson",
    "Flanders",
    "VanHouten",
    "Muntz",
    "Wiggum",
    "Nahasapeemapetilon",
    "Szyslak",
    "Leonard",
    "Carlson",
    "Gumble",
    "Krustofsky",
    "Burns",
    "Smithers",
    "Skinner",
    "Krabappel",
    "Bouvier",
    "Abe",
    "Lovejoy",
    "Hibbert",
    "Quimby",
    "Chalmers",
    "Frink",
    "Wolfcastle",
    "Otter",
    "Terwilliger",
    "McClure",
    "Brockman",
    "Prince",
    "Zoidberg",
    "Boimler",
    "Mariner",
    "Tendi",
    "Rutherford",
    "Freeman",
    "Shaxs",
    "Data",
    "Lwaxana",
    "Ro",
    "Yar",
    "Pulaski",
    "Wesley",
    "Shran",
    "Morn",
    "Garak",
    "Dukat",
    "Weyoun",
    "Romulan",
    "Cardassia",
    "Vulcan",
    "Andorian",
    "Bajoran",
    "Ferengi",
    "Hansen",
    "Rios",
    "Musiker",
    "Jurati",
    "Soong",
    "Noonian",
    "Talos",
    "Kelvin",
    "Daystrom",
    "Shelby",
    "Paris",
    "Locarno",
    "Sybok",
    "Vina",
]

STREET_POOL: List[str] = [
    "Baker Street",
    "Fremont Avenue",
    "King Street",
    "Elm Road",
    "Maple Avenue",
    "Sunset Boulevard",
    "Liberty Lane",
    "Harbor Street",
    "Riverside Drive",
    "Hillcrest Road",
    "Cedar Street",
    "Rosewood Avenue",
    "Park Lane",
    "Oak Street",
    "Pine Road",
    "Lakeview Street",
    "Willow Lane",
    "Market Street",
    "Broadway",
    "Station Road",
    "Victoria Street",
    "Church Lane",
    "Queens Road",
    "Mill Road",
    "Highfield Avenue",
    "Garden Street",
    "River Road",
    "Forest Avenue",
    "Meadow Lane",
    "School Street",
    "Bridge Street",
    "Canal Road",
    "Hill Street",
    "North Avenue",
    "South Street",
    "East Road",
    "West Lane",
    "New Street",
    "Old Mill Lane",
    "Waterfront Drive",
    "Seaview Road",
    "Orchard Street",
    "Valley Road",
    "Mountain View Drive",
    "Railway Street",
    "Airport Road",
    "College Avenue",
    "Museum Street",
    "Theater Lane",
    "Castle Road",
    "Beacon Street",
    "Palm Avenue",
    "Coral Road",
    "Harvest Lane",
    "Golden Gate Way",
    "Silver Street",
    "Copper Lane",
    "Pearl Avenue",
    "Granite Road",
    "Summit Street",
]

COUNTRY_POOL: List[str] = [
    "Iceland",
    "Canada",
    "Japan",
    "Chile",
    "Portugal",
    "New Zealand",
    "South Africa",
    "Finland",
    "Uruguay",
    "Slovenia",
    "Ireland",
    "Estonia",
    "Peru",
    "Vietnam",
    "Morocco",
    "Norway",
    "Argentina",
    "Croatia",
    "Kenya",
    "Thailand",
]

ID_REGEX = re.compile(r"\b\d+-[A-Z0-9]{3,8}-[A-Z0-9]{4,16}\b")
EMAIL_REGEX = re.compile(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b")
NAME_SPLIT_REGEX = re.compile(r"\s+")


@dataclass
class MappingStore:
    first_name_map: Dict[str, str]
    last_name_map: Dict[str, str]
    full_name_map: Dict[str, str]
    street_map: Dict[str, str]
    email_map: Dict[str, str]
    phone_map: Dict[str, str]
    postal_code_map: Dict[str, str]
    country_map: Dict[str, str]
    date_map: Dict[str, str]
    id_map: Dict[str, str]

    @classmethod
    def from_json(cls, data: Dict[str, Dict[str, str]]) -> "MappingStore":
        return cls(
            first_name_map=data.get("first_name_map", {}),
            last_name_map=data.get("last_name_map", {}),
            full_name_map=data.get("full_name_map", {}),
            street_map=data.get("street_map", {}),
            email_map=data.get("email_map", {}),
            phone_map=data.get("phone_map", {}),
            postal_code_map=data.get("postal_code_map", {}),
            country_map=data.get("country_map", {}),
            date_map=data.get("date_map", {}),
            id_map=data.get("id_map", {}),
        )

    def to_json(self) -> Dict[str, Dict[str, str]]:
        return {
            "first_name_map": self.first_name_map,
            "last_name_map": self.last_name_map,
            "full_name_map": self.full_name_map,
            "street_map": self.street_map,
            "email_map": self.email_map,
            "phone_map": self.phone_map,
            "postal_code_map": self.postal_code_map,
            "country_map": self.country_map,
            "date_map": self.date_map,
            "id_map": self.id_map,
        }


def normalize_header(value: str) -> str:
    return value.strip().lower()


def random_alnum(length: int, rng: random.Random) -> str:
    chars = string.ascii_uppercase + string.digits
    return "".join(rng.choice(chars) for _ in range(length))


def random_digits(length: int, rng: random.Random) -> str:
    if length <= 0:
        return ""
    head = rng.choice(string.digits[1:])
    tail = "".join(rng.choice(string.digits) for _ in range(length - 1))
    return head + tail


def random_like_segment(segment: str, rng: random.Random) -> str:
    out = []
    for ch in segment:
        if ch.isdigit():
            out.append(rng.choice(string.digits))
        elif ch.isalpha():
            out.append(rng.choice(string.ascii_uppercase))
        else:
            out.append(ch)
    return "".join(out)


def randomize_digits_by_pattern(value: str, rng: random.Random) -> str:
    chars: List[str] = []
    for ch in value:
        if ch.isdigit():
            chars.append(rng.choice(string.digits))
        else:
            chars.append(ch)
    return "".join(chars)


def anonymize_email_value(original: str, mapping: MappingStore, used_emails: Set[str], rng: random.Random) -> str:
    if original in mapping.email_map:
        return mapping.email_map[original]

    match = re.match(r"^([^@\s]+)@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})$", original.strip())
    if match:
        domain = match.group(2).lower()
    else:
        domain = "example.org"

    candidate = f"anon{rng.randint(100000, 999999)}@{domain}"
    while candidate in used_emails:
        candidate = f"anon{rng.randint(100000, 999999)}@{domain}"

    mapping.email_map[original] = candidate
    used_emails.add(candidate)
    return candidate


def anonymize_phone_value(original: str, mapping: MappingStore, used_phones: Set[str], rng: random.Random) -> str:
    if original in mapping.phone_map:
        return mapping.phone_map[original]

    candidate = randomize_digits_by_pattern(original, rng)
    while candidate in used_phones:
        candidate = randomize_digits_by_pattern(original, rng)

    mapping.phone_map[original] = candidate
    used_phones.add(candidate)
    return candidate


def anonymize_postal_code_value(original: str, mapping: MappingStore, used_postal_codes: Set[str], rng: random.Random) -> str:
    if original in mapping.postal_code_map:
        return mapping.postal_code_map[original]

    digits = "".join(ch for ch in original if ch.isdigit())
    if digits:
        replacement_digits = "".join(rng.choice(string.digits) for _ in range(len(digits)))
        di = 0
        chars: List[str] = []
        for ch in original:
            if ch.isdigit():
                chars.append(replacement_digits[di])
                di += 1
            else:
                chars.append(ch)
        candidate = "".join(chars)
    else:
        candidate = randomize_digits_by_pattern(original, rng)

    while candidate in used_postal_codes:
        candidate = randomize_digits_by_pattern(candidate, rng)

    mapping.postal_code_map[original] = candidate
    used_postal_codes.add(candidate)
    return candidate


def anonymize_country_value(original: str, mapping: MappingStore, used_countries: Set[str], rng: random.Random) -> str:
    return choose_mapped(original, mapping.country_map, COUNTRY_POOL, used_countries, rng)


def shift_year(dt: date, year_delta: int) -> date:
    new_year = dt.year + year_delta
    if new_year < 1900:
        new_year = 1900
    if new_year > 2099:
        new_year = 2099
    try:
        return dt.replace(year=new_year)
    except ValueError:
        # Handle leap day by moving to Feb 28.
        return dt.replace(year=new_year, day=28)


def anonymize_date_string(original: str, year_delta: int) -> str:
    patterns = [
        (r"^(\d{1,2}\.\d{1,2}\.\d{4})$", "%d.%m.%Y"),
        (r"^(\d{4}-\d{2}-\d{2})$", "%Y-%m-%d"),
        (r"^(\d{1,2}/\d{1,2}/\d{4})$", "%d/%m/%Y"),
    ]
    for pattern, fmt in patterns:
        if not re.match(pattern, original.strip()):
            continue
        try:
            dt = datetime.strptime(original.strip(), fmt).date()
            shifted = shift_year(dt, year_delta)
            return shifted.strftime(fmt)
        except ValueError:
            return original
    return original


def anonymize_date_value(original: object, date_kind: str, mapping: MappingStore, rng: random.Random) -> object:
    today = date.today()
    key = f"{date_kind}|{repr(original)}"
    if key in mapping.date_map:
        mapped = mapping.date_map[key]
        try:
            if isinstance(original, datetime):
                result = datetime.fromisoformat(mapped)
                if result.date() > today:
                    # Cached value is in future, regenerate
                    del mapping.date_map[key]
                else:
                    return result
            elif isinstance(original, date):
                result = date.fromisoformat(mapped)
                if result > today:
                    # Cached value is in future, regenerate
                    del mapping.date_map[key]
                else:
                    return result
            elif isinstance(original, int):
                return int(mapped)
            elif isinstance(original, float):
                return float(mapped)
            else:
                return mapped
        except (ValueError, AttributeError):
            pass

    year_delta = rng.randint(-25, 0) if date_kind == "birth_date" else rng.randint(-10, 0)
    if year_delta == 0:
        year_delta = -1

    if isinstance(original, datetime):
        shifted = datetime.combine(shift_year(original.date(), year_delta), original.time())
        # Ensure not in future
        if shifted.date() > today:
            shifted = datetime.combine(today, shifted.time())
        mapping.date_map[key] = shifted.isoformat()
        return shifted
    if isinstance(original, date):
        shifted = shift_year(original, year_delta)
        # Ensure not in future
        if shifted > today:
            shifted = today
        mapping.date_map[key] = shifted.isoformat()
        return shifted
    if isinstance(original, int):
        shifted_int = original + (365 * year_delta)
        mapping.date_map[key] = str(shifted_int)
        return shifted_int
    if isinstance(original, float):
        shifted_float = original + (365.0 * year_delta)
        mapping.date_map[key] = str(shifted_float)
        return shifted_float
    if isinstance(original, str):
        shifted_str = anonymize_date_string(original, year_delta)
        # Ensure string date is not in future
        try:
            for pattern, fmt in [
                (r"^(\d{1,2}\.\d{1,2}\.\d{4})$", "%d.%m.%Y"),
                (r"^(\d{4}-\d{2}-\d{2})$", "%Y-%m-%d"),
                (r"^(\d{1,2}/\d{1,2}/\d{4})$", "%d/%m/%Y"),
            ]:
                if re.match(pattern, shifted_str.strip()):
                    parsed_date = datetime.strptime(shifted_str.strip(), fmt).date()
                    if parsed_date > today:
                        shifted_str = today.strftime(fmt)
                    break
        except (ValueError, AttributeError):
            pass
        mapping.date_map[key] = shifted_str
        return shifted_str

    return original


def anonymize_id_value(original: str, mapping: MappingStore, used_ids: Set[str], rng: random.Random) -> str:
    if original in mapping.id_map:
        return mapping.id_map[original]

    parts = original.split("-")
    if len(parts) != 3:
        candidate = f"{random_digits(1, rng)}-{random_alnum(4, rng)}-{random_alnum(6, rng)}"
    else:
        p1, p2, p3 = parts
        c1 = random_digits(len(p1), rng) if p1.isdigit() else random_like_segment(p1, rng)
        c2 = random_like_segment(p2, rng)
        c3 = random_like_segment(p3, rng)
        candidate = f"{c1}-{c2}-{c3}"

    while candidate in used_ids:
        p1, p2, p3 = candidate.split("-")
        candidate = f"{random_like_segment(p1, rng)}-{random_like_segment(p2, rng)}-{random_like_segment(p3, rng)}"

    mapping.id_map[original] = candidate
    used_ids.add(candidate)
    return candidate


def choose_mapped(
    original: str,
    mapping_dict: Dict[str, str],
    pool: List[str],
    used_values: Set[str],
    rng: random.Random,
) -> str:
    if original in mapping_dict:
        return mapping_dict[original]

    candidates = [x for x in pool if x not in used_values]
    if not candidates:
        # Fallback if the pool is exhausted.
        candidate = f"{rng.choice(pool)}{rng.randint(10, 999)}"
    else:
        candidate = rng.choice(candidates)

    mapping_dict[original] = candidate
    used_values.add(candidate)
    return candidate


def anonymize_full_name(value: str, mapping: MappingStore, used_first: Set[str], used_last: Set[str], rng: random.Random) -> str:
    if value in mapping.full_name_map:
        return mapping.full_name_map[value]

    parts = [p for p in NAME_SPLIT_REGEX.split(value.strip()) if p]
    if not parts:
        return value

    if len(parts) == 1:
        anon = choose_mapped(parts[0], mapping.first_name_map, FIRST_NAME_POOL, used_first, rng)
    else:
        anon_first = choose_mapped(parts[0], mapping.first_name_map, FIRST_NAME_POOL, used_first, rng)
        anon_last = choose_mapped(parts[-1], mapping.last_name_map, LAST_NAME_POOL, used_last, rng)
        middle = parts[1:-1]
        anon_middle = [choose_mapped(p, mapping.first_name_map, FIRST_NAME_POOL, used_first, rng) for p in middle]
        anon = " ".join([anon_first, *anon_middle, anon_last])

    mapping.full_name_map[value] = anon
    return anon


def anonymize_street(value: str, mapping: MappingStore, used_streets: Set[str], rng: random.Random) -> str:
    if value in mapping.street_map:
        return mapping.street_map[value]

    # Keep house number and tail, replace only street name at the beginning.
    match = re.match(r"^\s*([^\d,]+?)(\s+\d+[A-Za-z0-9/\-]*.*)?$", value.strip())
    if match:
        street_name = match.group(1).strip()
        suffix = match.group(2) or ""
    else:
        street_name = value.strip()
        suffix = ""

    anon_street = choose_mapped(street_name, mapping.street_map, STREET_POOL, used_streets, rng)
    result = (anon_street + suffix).strip()
    mapping.street_map[value] = result
    return result


def replace_embedded_ids(text: str, mapping: MappingStore, used_ids: Set[str], rng: random.Random) -> str:
    def repl(match: re.Match[str]) -> str:
        return anonymize_id_value(match.group(0), mapping, used_ids, rng)

    return ID_REGEX.sub(repl, text)


def replace_embedded_emails(text: str, mapping: MappingStore, used_emails: Set[str], rng: random.Random) -> str:
    def repl(match: re.Match[str]) -> str:
        return anonymize_email_value(match.group(0), mapping, used_emails, rng)

    return EMAIL_REGEX.sub(repl, text)


def classify_column(header: str) -> str:
    h = normalize_header(header)
    if not h:
        return "other"

    if "id" in h:
        return "id"
    if "e-mail" in h or "email" in h:
        return "email"
    if "telefon" in h or "handy" in h or "fax" in h:
        return "phone"
    if h == "geburtsdatum":
        return "birth_date"
    if h in {"eintrittsdatum", "beginn mitgliedschaft"}:
        return "entry_date"
    if h == "plz":
        return "postal_code"
    if h == "land" or "country" in h:
        return "country"
    if h in {"vorname", "zweiter name", "spitzname"}:
        return "first_name"
    if h == "nachname":
        return "last_name"
    if h in {"kontakt 1", "kontakt 2", "erstellt von", "aktualisiert von"}:
        return "full_name"
    if "strasse" in h or "adresse" in h or "street" in h:
        return "street"
    return "other"


def load_mapping(path: Path) -> MappingStore:
    if not path.exists():
        return MappingStore({}, {}, {}, {}, {}, {}, {}, {}, {}, {})
    with path.open("r", encoding="utf-8") as f:
        return MappingStore.from_json(json.load(f))


def save_mapping(path: Path, mapping: MappingStore) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(mapping.to_json(), f, indent=2, ensure_ascii=False)


def anonymize_workbook(input_file: Path, output_file: Path, mapping_file: Path, seed: int) -> None:
    rng = random.Random(seed)
    mapping = load_mapping(mapping_file)

    used_first = set(mapping.first_name_map.values())
    used_last = set(mapping.last_name_map.values())
    used_streets = set(mapping.street_map.values())
    used_emails = set(mapping.email_map.values())
    used_phones = set(mapping.phone_map.values())
    used_postal_codes = set(mapping.postal_code_map.values())
    used_countries = set(mapping.country_map.values())
    used_ids = set(mapping.id_map.values())

    wb = load_workbook(input_file)

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue

        header_map: Dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if isinstance(value, str):
                header_map[col] = classify_column(value)
            else:
                header_map[col] = "other"

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                column_type = header_map.get(col, "other")

                if value is None:
                    continue

                if isinstance(value, str):
                    if not value.strip():
                        continue
                    value = replace_embedded_emails(value, mapping, used_emails, rng)
                    value = replace_embedded_ids(value, mapping, used_ids, rng)

                if column_type == "id" and isinstance(value, str) and ID_REGEX.fullmatch(value):
                    value = anonymize_id_value(value, mapping, used_ids, rng)
                elif column_type == "email" and isinstance(value, str) and EMAIL_REGEX.fullmatch(value):
                    value = anonymize_email_value(value, mapping, used_emails, rng)
                elif column_type == "phone":
                    value = anonymize_phone_value(str(value), mapping, used_phones, rng)
                elif column_type in {"birth_date", "entry_date"}:
                    value = anonymize_date_value(value, column_type, mapping, rng)
                elif column_type == "postal_code":
                    value = anonymize_postal_code_value(str(value), mapping, used_postal_codes, rng)
                elif column_type == "country":
                    value = anonymize_country_value(str(value), mapping, used_countries, rng)
                elif column_type == "first_name":
                    value = choose_mapped(value, mapping.first_name_map, FIRST_NAME_POOL, used_first, rng)
                elif column_type == "last_name":
                    value = choose_mapped(value, mapping.last_name_map, LAST_NAME_POOL, used_last, rng)
                elif column_type == "full_name":
                    # Keep role prefixes like "Dr." if present by anonymizing only name-like tokens.
                    value = anonymize_full_name(value, mapping, used_first, used_last, rng)
                elif column_type == "street":
                    value = anonymize_street(value, mapping, used_streets, rng)

                cell.value = value

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    save_mapping(mapping_file, mapping)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Anonymize names, streets, emails, phone numbers, dates, postal codes, countries and IDs in an Excel file."
    )
    parser.add_argument("input", type=Path, help="Input .xlsx file")
    parser.add_argument("output", type=Path, help="Output anonymized .xlsx file")
    parser.add_argument(
        "--mapping",
        type=Path,
        default=Path("testdata/anonymization_mapping.json"),
        help="JSON mapping file reused across runs",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=20260421,
        help="Random seed for reproducibility of new mappings",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    anonymize_workbook(args.input, args.output, args.mapping, args.seed)
    print(f"Anonymized workbook written to: {args.output}")
    print(f"Mapping file written to: {args.mapping}")


if __name__ == "__main__":
    main()
